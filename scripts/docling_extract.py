import sys
import json
import os
import argparse
import csv
import tempfile
import subprocess
from typing import Iterable, Optional

# Phase 1 (Canonical Reader Wiring):
# This is the SAME existing extraction stack already wired through
# lib/protocol-intake/extractors/document-extraction-adapter.ts.
# It is enhanced (not replaced) to additionally expose:
#   - full_text       (document text so downstream section/candidate extractors can run)
#   - tables[].page_no (page provenance when the reader exposes it)
#   - extraction_method / page_count (provenance preservation)
# No new reader is introduced; PDF/DOCX use docling, XLSX uses openpyxl, CSV uses stdlib.

DOCLING_CHUNK_PAGES = int(os.environ.get("PROTOCOL_INTAKE_DOCLING_CHUNK_PAGES", "8"))
TEXT_SUCCESS_THRESHOLD = int(os.environ.get("PROTOCOL_INTAKE_TEXT_SUCCESS_THRESHOLD", "2000"))


def _page_ranges(page_count: int, chunk_size: int) -> Iterable[tuple[int, int]]:
    start = 1
    while start <= page_count:
        end = min(page_count, start + chunk_size - 1)
        yield start, end
        start = end + 1


def _subset_pdf(file_path: str, start_page: int, end_page: int) -> str:
    import fitz

    source = fitz.open(file_path)
    subset = fitz.open()
    subset.insert_pdf(source, from_page=start_page - 1, to_page=end_page - 1)
    handle = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    try:
        subset.save(handle.name)
    finally:
        handle.close()
        source.close()
        subset.close()
    return handle.name


def _table_to_markdown(grid):
    rows = []
    for row in grid or []:
      if row is None:
          continue
      rows.append([str(cell).replace("\n", " ").strip() if cell is not None else "" for cell in row])
    if not rows:
        return ""
    header = rows[0]
    lines = [" | ".join(header), " | ".join(["---"] * len(header))]
    for row in rows[1:]:
        lines.append(" | ".join(row))
    return "\n".join(lines)


def _pypdfium2_extract(file_path, page_numbers: Optional[Iterable[int]] = None):
    import pypdfium2

    doc = pypdfium2.PdfDocument(str(file_path))
    try:
        page_count = len(doc)
        if page_numbers is None:
            pages = list(range(1, page_count + 1))
        else:
            pages = [p for p in page_numbers if 1 <= p <= page_count]

        text_parts = []
        for page_no in pages:
            page = doc[page_no - 1]
            textpage = page.get_textpage()
            try:
                page_text = textpage.get_text_range() or ""
            finally:
                try:
                    textpage.close()
                except Exception:
                    pass
            page_text = page_text.strip()
            if page_text:
                text_parts.append(f"\n\n[Page {page_no}]\n{page_text}")
        full_text = "\n".join(text_parts).strip()
        return full_text, [], page_count, "pypdfium2"
    finally:
        try:
            doc.close()
        except Exception:
            pass


def _docling_extract(file_path):
    """PDF / DOCX via docling. Returns (full_text, tables, page_count)."""
    from docling.document_converter import DocumentConverter

    converter = DocumentConverter()
    result = converter.convert(file_path)
    document = result.document

    try:
        full_text = document.export_to_markdown()
    except Exception:
        full_text = ""

    tables = []
    for table in getattr(document, "tables", []) or []:
        page_no = None
        try:
            prov = getattr(table, "prov", None) or []
            if prov:
                page_no = getattr(prov[0], "page_no", None)
        except Exception:
            page_no = None
        try:
            tables.append(
                {
                    "table_html": table.export_to_html(),
                    "table_markdown": table.export_to_markdown(),
                    "page_no": page_no,
                }
            )
        except Exception:
            continue

    page_count = None
    try:
        pages = getattr(document, "pages", None)
        if pages is not None:
            page_count = len(pages)
    except Exception:
        page_count = None

    return full_text, tables, page_count, "docling"


def _run_docling_child(file_path: str):
    cmd = [sys.executable, os.path.abspath(__file__), "--docling-only", file_path]
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
      raise RuntimeError(
          (proc.stderr or proc.stdout or f"docling child failed with code {proc.returncode}").strip()
      )
    try:
        payload = json.loads(proc.stdout)
    except Exception as exc:
        raise RuntimeError(f"docling child returned invalid JSON: {exc}") from exc
    if payload.get("error"):
        raise RuntimeError(str(payload.get("error")))
    return payload


def _targeted_page_numbers(file_path, page_count):
    import pypdfium2

    doc = pypdfium2.PdfDocument(str(file_path))
    try:
        keywords = [
            "schedule of activities",
            "schedule of events",
            "soa",
            "screening",
            "baseline",
            "follow-up",
            "follow up",
            "treatment",
            "day 0",
            "day 3",
            "day 5",
            "day 9",
            "day 21",
            "index patient",
            "household contact",
            "remote",
            "home",
            "phone",
            "swab",
            "sample",
            "adverse event",
            "concomitant",
            "questionnaire",
        ]
        hits = set()
        total_pages = min(page_count, len(doc))
        for page_no in range(1, total_pages + 1):
            page = doc[page_no - 1]
            textpage = page.get_textpage()
            try:
                text = (textpage.get_text_range() or "").lower()
            finally:
                try:
                    textpage.close()
                except Exception:
                    pass
            if any(k in text for k in keywords):
                hits.add(page_no)
                if page_no > 1:
                    hits.add(page_no - 1)
                if page_no < page_count:
                    hits.add(page_no + 1)
        return sorted(p for p in hits if 1 <= p <= page_count)
    finally:
        try:
            doc.close()
        except Exception:
            pass


def _excel_extract(file_path):
    """XLSX via openpyxl. Returns (full_text, tables, page_count)."""
    import openpyxl

    wb = openpyxl.load_workbook(file_path, data_only=True)
    tables = []
    text_parts = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append([str(cell) if cell is not None else "" for cell in row])

        text_parts.append("Sheet: %s" % sheet_name)
        html = "<table>"
        md_lines = []
        for row in data:
            html += "<tr>" + "".join("<td>%s</td>" % c for c in row) + "</tr>"
            md_lines.append(" | ".join(row))
        html += "</table>"
        text_parts.append("\n".join(md_lines))

        tables.append(
            {
                "table_html": html,
                "table_markdown": "\n".join(md_lines),
                "page_no": None,
            }
        )

    return "\n\n".join(text_parts), tables, None, "excel"


def _csv_extract(file_path):
    """CSV via stdlib. Returns (full_text, tables, page_count)."""
    rows = []
    with open(file_path, encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.reader(f)
        for row in reader:
            rows.append([str(c) for c in row])

    html = "<table>"
    md_lines = []
    for row in rows:
        html += "<tr>" + "".join("<td>%s</td>" % c for c in row) + "</tr>"
        md_lines.append(" | ".join(row))
    html += "</table>"

    full_text = "\n".join(md_lines)
    tables = [{"table_html": html, "table_markdown": full_text, "page_no": None}]
    return full_text, tables, None, "csv"


def main():
    parser = argparse.ArgumentParser(
        description="Extract text + tables from PDF/DOCX/XLSX/CSV using docling/openpyxl/stdlib."
    )
    parser.add_argument("file_path", help="Path to the input file")
    parser.add_argument("--docling-only", action="store_true", help="Run the docling reader only and emit JSON.")
    args = parser.parse_args()

    file_path = args.file_path
    if not os.path.exists(file_path):
        print(json.dumps({"error": "File not found: %s" % file_path, "attempted_reader": "none", "failure_stage": "input"}))
        sys.exit(1)

    ext = os.path.splitext(file_path)[1].lower()
    try:
        if args.docling_only:
            full_text, tables, page_count, method = _docling_extract(file_path)
            print(
                json.dumps(
                    {
                        "full_text": full_text,
                        "tables": tables,
                        "page_count": page_count,
                        "extraction_method": method,
                        "attempted_reader": method,
                    }
                )
            )
            return

        if ext in [".pdf", ".docx"]:
            file_page_count = None
            if ext == ".pdf":
                try:
                    import fitz
                    with fitz.open(file_path) as doc:
                        file_page_count = doc.page_count
                except Exception:
                    file_page_count = None

            attempts = []
            full_text = ""
            tables = []
            page_count = file_page_count
            method = ""

            try:
                child = _run_docling_child(file_path)
                full_text = child.get("full_text") or ""
                tables = child.get("tables") or []
                page_count = child.get("page_count", page_count)
                method = child.get("extraction_method") or "docling"
                attempts.append(method)
                if len(full_text or "") < TEXT_SUCCESS_THRESHOLD:
                    raise RuntimeError("docling returned insufficient text")
            except Exception as first_error:
                attempts.append("docling_failed")
                try:
                    targeted_pages = []
                    if file_page_count and file_page_count > 1:
                        targeted_pages = _targeted_page_numbers(file_path, file_page_count)
                    if targeted_pages:
                        targeted_text, targeted_tables, targeted_page_count, method = _pypdfium2_extract(file_path, targeted_pages)
                        attempts.append(f"targeted:{','.join(map(str, targeted_pages[:25]))}")
                        attempts.append(method)
                    else:
                        targeted_text, targeted_tables, targeted_page_count, method = "", [], file_page_count, "pypdfium2"

                    if len(targeted_text or "") >= TEXT_SUCCESS_THRESHOLD:
                        full_text, tables, page_count = targeted_text, targeted_tables, targeted_page_count
                    else:
                        full_text, tables, page_count, method = _pypdfium2_extract(file_path)
                        attempts.append(method)
                except Exception as fallback_error:
                    print(
                        json.dumps(
                            {
                                "error": str(fallback_error),
                                "attempted_reader": "docling+pypdfium2",
                                "failure_stage": "pdf_extraction",
                                "partial_text": "",
                                "page_range": None,
                            }
                        )
                    )
                    sys.exit(1)

            print(
                json.dumps(
                    {
                        "full_text": full_text,
                        "tables": tables,
                        "page_count": page_count,
                        "extraction_method": method,
                        "attempted_reader": attempts[-1] if attempts else method,
                        "attempted_readers": attempts,
                    }
                )
            )
            return
        elif ext in [".xlsx"]:
            full_text, tables, page_count, method = _excel_extract(file_path)
        elif ext in [".csv"]:
            full_text, tables, page_count, method = _csv_extract(file_path)
        else:
            print(json.dumps({"error": "Unsupported file extension: %s" % ext}))
            sys.exit(1)

        print(
            json.dumps(
                {
                    "full_text": full_text,
                    "tables": tables,
                    "page_count": page_count,
                    "extraction_method": method,
                    "attempted_reader": method,
                }
            )
        )
    except Exception as e:
        print(json.dumps({"error": str(e), "attempted_reader": "unknown", "failure_stage": "script"}))
        sys.exit(1)


if __name__ == "__main__":
    main()
