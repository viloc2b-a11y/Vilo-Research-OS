"""Document readers — optional PyMuPDF/pdfplumber/docx/openpyxl with stdlib fallbacks."""
from __future__ import annotations

import csv
import io
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

DOCX_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}


def file_type_for(path: Path) -> str:
    ext = path.suffix.lower()
    if ext == ".pdf":
        return "pdf"
    if ext in {".docx"}:
        return "docx"
    if ext in {".xlsx", ".xls"}:
        return "excel"
    if ext in {".csv"}:
        return "csv"
    if ext in {".txt", ".md"}:
        return "plain_text"
    return "unknown"


def read_plain_text(path: Path) -> dict[str, Any]:
    text = path.read_text(encoding="utf-8", errors="replace")
    return {
        "file_name": path.name,
        "file_type": "plain_text",
        "pages": [{"page_number": 1, "text": text}],
        "sheets": [],
        "sections": [],
        "tables": [],
        "footnotes": [],
    }


def read_csv_as_sheet(path: Path) -> dict[str, Any]:
    rows: list[dict[str, str]] = []
    with path.open(encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for row in reader:
            rows.append({k: (row.get(k) or "").strip() for k in headers})
    return {
        "file_name": path.name,
        "file_type": "csv",
        "pages": [],
        "sheets": [{"sheet_name": path.stem, "headers": list(headers), "rows": rows}],
        "sections": [],
        "tables": [
            {
                "sheet_name": path.stem,
                "row_start": 1,
                "row_end": len(rows) + 1,
                "column_headers": list(headers),
                "rows": rows,
            }
        ],
        "footnotes": [],
    }


def _read_docx_stdlib(path: Path) -> str:
    with zipfile.ZipFile(path) as zf:
        xml = zf.read("word/document.xml")
    root = ET.fromstring(xml)
    paragraphs: list[str] = []
    for p in root.findall(".//w:p", DOCX_NS):
        texts = [t.text for t in p.findall(".//w:t", DOCX_NS) if t.text]
        if texts:
            paragraphs.append("".join(texts))
    return "\n".join(paragraphs)


def read_docx(path: Path) -> dict[str, Any]:
    text = ""
    try:
        import docx  # type: ignore

        doc = docx.Document(str(path))
        text = "\n".join(p.text for p in doc.paragraphs)
    except Exception:
        text = _read_docx_stdlib(path)
    return {
        "file_name": path.name,
        "file_type": "docx",
        "pages": [{"page_number": 1, "text": text}],
        "sheets": [],
        "sections": [],
        "tables": [],
        "footnotes": [],
    }


def read_pdf(path: Path) -> dict[str, Any]:
    pages: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    footnotes: list[dict[str, Any]] = []

    try:
        import fitz  # PyMuPDF  # type: ignore

        doc = fitz.open(str(path))
        for i, page in enumerate(doc, start=1):
            pages.append({"page_number": i, "text": page.get_text("text")})
    except Exception:
        pages = [{"page_number": 1, "text": path.read_bytes()[:0].decode("utf-8", errors="replace")}]

    try:
        import pdfplumber  # type: ignore

        with pdfplumber.open(str(path)) as pdf:
            for i, page in enumerate(pdf.pages, start=1):
                for ti, table in enumerate(page.extract_tables() or []):
                    if not table:
                        continue
                    headers = [str(c or "").strip() for c in table[0]]
                    rows = [
                        {headers[j]: str(row[j] or "").strip() for j in range(len(headers))}
                        for row in table[1:]
                        if row
                    ]
                    tables.append(
                        {
                            "page_number": i,
                            "table_index": ti,
                            "column_headers": headers,
                            "rows": rows,
                        }
                    )
    except Exception:
        pass

    if not pages or not any(p.get("text") for p in pages):
        pages = [{"page_number": 1, "text": ""}]

    return {
        "file_name": path.name,
        "file_type": "pdf",
        "pages": pages,
        "sheets": [],
        "sections": [],
        "tables": tables,
        "footnotes": footnotes,
    }


def read_excel(path: Path) -> dict[str, Any]:
    sheets: list[dict[str, Any]] = []
    tables: list[dict[str, Any]] = []
    try:
        import openpyxl  # type: ignore

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        for name in sorted(wb.sheetnames):
            ws = wb[name]
            rows_raw = list(ws.iter_rows(values_only=True))
            if not rows_raw:
                continue
            headers = [str(c or "").strip() for c in rows_raw[0]]
            rows = []
            for r in rows_raw[1:]:
                rows.append({headers[i]: str(r[i] or "").strip() for i in range(len(headers))})
            sheets.append({"sheet_name": name, "headers": headers, "rows": rows})
            tables.append(
                {
                    "sheet_name": name,
                    "row_start": 1,
                    "row_end": len(rows) + 1,
                    "column_headers": headers,
                    "rows": rows,
                }
            )
        wb.close()
        return {
            "file_name": path.name,
            "file_type": "excel",
            "pages": [],
            "sheets": sheets,
            "tables": tables,
            "footnotes": [],
        }
    except Exception:
        pass

    try:
        import pandas as pd  # type: ignore

        xls = pd.read_excel(str(path), sheet_name=None)
        for name in sorted(xls.keys()):
            df = xls[name].fillna("")
            headers = [str(c) for c in df.columns]
            rows = [dict(zip(headers, (str(v) for v in row))) for row in df.to_dict(orient="records")]
            sheets.append({"sheet_name": str(name), "headers": headers, "rows": rows})
            tables.append(
                {
                    "sheet_name": str(name),
                    "row_start": 1,
                    "row_end": len(rows) + 1,
                    "column_headers": headers,
                    "rows": rows,
                }
            )
    except Exception:
        return read_plain_text(path)

    return {
        "file_name": path.name,
        "file_type": "excel",
        "pages": [],
        "sheets": sheets,
        "tables": tables,
        "footnotes": [],
    }


def read_document(path: Path) -> dict[str, Any]:
    ft = file_type_for(path)
    if ft == "pdf":
        return read_pdf(path)
    if ft == "docx":
        return read_docx(path)
    if ft == "excel":
        return read_excel(path)
    if ft == "csv":
        return read_csv_as_sheet(path)
    return read_plain_text(path)
